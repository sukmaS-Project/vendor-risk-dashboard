"""
=============================================================
  PREDICT.PY — KLASIFIKASI & PREDIKSI RISIKO VENDOR
  Penelitian: Klasifikasi Risiko Vendor TAD (2023-2025)
  Model: Random Forest Eks.4 (lag + class weight)
         CV Macro F1 = 0.9033
=============================================================
OUTPUT YANG DIHASILKAN (dari 1 input):

  OUTPUT 1 — Monitoring Bulan Ini
  Klasifikasi risiko aktual berdasarkan data yang diinput.
  Dihitung dari formula skor risiko expert judgment.
  Tidak menggunakan model ML.

  OUTPUT 2 — Prediksi Bulan Depan
  Prediksi risiko bulan berikutnya menggunakan model RF Eks.4.
  Menggunakan carry-forward assumption: fitur non-lag bulan
  depan diestimasi dari nilai aktual bulan ini, sedangkan
  lag bulan depan = nilai aktual bulan ini.

CARA PENGGUNAAN SETIAP BULAN:
  1. Buka Data_Input_Bulan_Ini.xlsx
  2. Isi data 5 vendor aktif (sel kuning) — bulan berjalan
  3. Simpan file
  4. Jalankan: python predict.py
  5. Hasil tersimpan di output_prediksi/

VENDOR AKTIF: PT CSMP, PT LJ, PT MS, PT PMS, PT UJP
=============================================================
"""

import pandas as pd
import numpy as np
import pickle
import warnings
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment,
                              Border, Side)
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

# ─────────────────────────────────────────────────────────────
# KONFIGURASI
# ─────────────────────────────────────────────────────────────
MODEL_PATH   = Path('output_7eksperimen/best_model.pkl')
HISTORY_FILE = 'Data.xlsx'
OUTPUT_DIR   = Path('output_prediksi')
OUTPUT_DIR.mkdir(exist_ok=True)

VENDORS_AKTIF = ['PT CSMP', 'PT LJ', 'PT MS', 'PT PMS', 'PT UJP']

BULAN_ORDER = [
    'JANUARI', 'FEBRUARI', 'MARET', 'APRIL', 'MEI', 'JUNI',
    'JULI', 'AGUSTUS', 'SEPTEMBER', 'OKTOBER', 'NOVEMBER', 'DESEMBER'
]

# Fitur non-lag (7 fitur) — diisi manual di Excel
FITUR_NON_LAG = [
    'Proportion Delay',
    'Proportion Gap',
    'Mean Delay (X01)',
    'Mean Delay (X02)',
    'Mean Delay (X04)',
    'Mean Gap (X11)',
    'Mean Gap (X14)',
]

# Fitur lag (6 fitur) — dihitung otomatis
FITUR_LAG = [
    'Lag Proportion Delay',
    'Lag Proportion Gap',
    'Lag Mean Delay (X02)',
    'Lag Mean Delay (X04)',
    'Lag  Mean Gap (X11)',
    'Lag Mean Gap (X14)',
]

# Mapping: kolom lag → kolom sumber di historis
LAG_SOURCE_MAP = {
    'Lag Proportion Delay' : 'Proportion Delay',
    'Lag Proportion Gap'   : 'Proportion Gap',
    'Lag Mean Delay (X02)' : 'Mean Delay (X02)',
    'Lag Mean Delay (X04)' : 'Mean Delay (X04)',
    'Lag  Mean Gap (X11)'  : 'Mean Gap (X11)',
    'Lag Mean Gap (X14)'   : 'Mean Gap (X14)',
}

ALL_FEATURES = FITUR_NON_LAG + FITUR_LAG

# Formula skor risiko expert judgment
BOBOT_DELAY = 0.4
BOBOT_GAP   = 0.6

WARNA = {
    'Patuh'  : {'bg': 'C6EFCE', 'font': '276221'},
    'Waspada': {'bg': 'FFEB9C', 'font': '9C5700'},
    'Kritis' : {'bg': 'FFC7CE', 'font': '9C0006'},
}

REKOMENDASI = {
    'Patuh'  : 'Pertahankan kinerja, monitoring rutin',
    'Waspada': 'Tingkatkan pemantauan, kirim peringatan tertulis',
    'Kritis' : 'Eskalasi segera ke manajemen, tindak lanjuti kontrak',
}


# ═══════════════════════════════════════════════════════════════
# FUNGSI UTILITAS
# ═══════════════════════════════════════════════════════════════

def bulan_berikutnya(bulan, tahun):
    """Hitung bulan dan tahun berikutnya."""
    idx = BULAN_ORDER.index(bulan.upper())
    if idx == 11:  # Desember -> Januari tahun berikutnya
        return 'JANUARI', tahun + 1
    return BULAN_ORDER[idx + 1], tahun


def hitung_skor_risiko(proportion_delay, proportion_gap):
    """
    Hitung skor risiko menggunakan formula expert judgment.
    Skor = (0.4 * Proportion_Delay + 0.6 * Proportion_Gap) * 100
    """
    return (BOBOT_DELAY * proportion_delay +
            BOBOT_GAP   * proportion_gap) * 100


def klasifikasi_dari_skor(skor):
    """Klasifikasi berdasarkan threshold skor risiko."""
    if skor == 0:
        return 'Patuh'
    elif skor <= 25:
        return 'Waspada'
    else:
        return 'Kritis'


def load_model():
    """Load model artifact dari file pkl."""
    if not MODEL_PATH.exists():
        raise FileNotFoundError(
            f"\n  File model tidak ditemukan: {MODEL_PATH}"
            "\n  Pastikan sudah menjalankan eksperimen_7.py."
        )
    with open(MODEL_PATH, 'rb') as f:
        return pickle.load(f)


def load_history():
    """Load dan siapkan data historis untuk kalkulasi lag."""
    if not Path(HISTORY_FILE).exists():
        raise FileNotFoundError(
            f"\n  File historis tidak ditemukan: {HISTORY_FILE}"
        )
    df = pd.read_excel(HISTORY_FILE)
    if 'MLag ean Delay (X04)' in df.columns:
        df = df.rename(
            columns={'MLag ean Delay (X04)': 'Lag Mean Delay (X04)'}
        )
    df['Bulan_sort'] = pd.Categorical(
        df['Bulan Tagihan'].str.upper(),
        categories=BULAN_ORDER, ordered=True
    )
    df = df.sort_values(
        ['Nama Vendor', 'Tahun Tagihan', 'Bulan_sort']
    ).reset_index(drop=True)
    df['periode_num'] = (
        df['Tahun Tagihan'] * 12 + df['Bulan_sort'].cat.codes
    )
    return df


def ambil_lag(df_history, vendor, bulan, tahun):
    """
    Ambil nilai lag dari data historis untuk vendor dan periode.
    Mencari baris dengan periode = (tahun*12 + idx_bulan) - 1
    """
    bulan_upper  = bulan.strip().upper()
    periode_ini  = tahun * 12 + BULAN_ORDER.index(bulan_upper)
    periode_prev = periode_ini - 1

    baris = df_history[
        (df_history['Nama Vendor'] == vendor) &
        (df_history['periode_num'] == periode_prev)
    ]

    if len(baris) == 0:
        return {c: 0.0 for c in FITUR_LAG}, False

    row = baris.iloc[0]
    return {
        lag_col: float(row.get(src_col, 0.0))
        for lag_col, src_col in LAG_SOURCE_MAP.items()
    }, True


def prediksi_dengan_model(artifact, fitur_row):
    """
    Jalankan prediksi model ML untuk satu baris fitur.
    Return: label prediksi, dict probabilitas
    """
    model     = artifact['model']
    scaler    = artifact['scaler']
    label_inv = artifact['label_inv']

    X        = np.array([[fitur_row[f] for f in ALL_FEATURES]])
    X_scaled = scaler.transform(X)

    pred_code  = model.predict(X_scaled)[0]
    pred_label = label_inv[int(pred_code)]

    if hasattr(model, 'predict_proba'):
        proba = model.predict_proba(X_scaled)[0]
        prob  = {'Patuh': proba[0],
                 'Waspada': proba[1],
                 'Kritis': proba[2]}
    else:
        prob = {'Patuh': 0, 'Waspada': 0, 'Kritis': 0}

    return pred_label, prob


# ═══════════════════════════════════════════════════════════════
# PROSES UTAMA — DUA OUTPUT
# ═══════════════════════════════════════════════════════════════

def proses_semua(artifact, df_history, df_input):
    """
    Proses setiap vendor dan hasilkan dua output:
      1. Klasifikasi bulan ini (formula skor risiko)
      2. Prediksi bulan depan (model ML + carry-forward)
    """
    hasil_monitoring = []
    hasil_prediksi   = []

    for _, row in df_input.iterrows():
        vendor = str(row['Nama Vendor']).strip()
        bulan  = str(row['Bulan Tagihan']).strip().upper()
        tahun  = int(row['Tahun Tagihan'])

        # Ambil nilai fitur non-lag bulan ini
        fitur_ini = {f: float(row.get(f, 0.0)) for f in FITUR_NON_LAG}

        # ── OUTPUT 1: KLASIFIKASI BULAN INI ──────────────────
        # Hitung skor risiko dari formula expert judgment
        skor   = hitung_skor_risiko(
            fitur_ini['Proportion Delay'],
            fitur_ini['Proportion Gap']
        )
        label_monitoring = klasifikasi_dari_skor(skor)

        # Ambil lag bulan ini (dari bulan sebelumnya di historis)
        lag_bulan_ini, lag_found = ambil_lag(
            df_history, vendor, bulan, tahun
        )

        # Jalankan model untuk monitoring (fitur ini + lag bulan lalu)
        fitur_monitoring = {**fitur_ini, **lag_bulan_ini}
        label_model_monitoring, prob_monitoring = prediksi_dengan_model(
            artifact, fitur_monitoring
        )

        hasil_monitoring.append({
            'Nama Vendor'         : vendor,
            'Bulan'               : bulan,
            'Tahun'               : tahun,
            'Skor Risiko'         : round(skor, 2),
            'Klasifikasi (Formula)': label_monitoring,
            'Klasifikasi (Model)' : label_model_monitoring,
            'Prob Patuh'          : prob_monitoring['Patuh'],
            'Prob Waspada'        : prob_monitoring['Waspada'],
            'Prob Kritis'         : prob_monitoring['Kritis'],
            'Probabilitas'        : (
                f"Patuh={prob_monitoring['Patuh']:.0%} | "
                f"Waspada={prob_monitoring['Waspada']:.0%} | "
                f"Kritis={prob_monitoring['Kritis']:.0%}"
            ),
            'Rekomendasi'         : REKOMENDASI.get(
                label_model_monitoring, '-'
            ),
            'Status Lag'          : (
                'lag dari bulan sebelumnya' if lag_found
                else 'lag=0 (bulan pertama)'
            ),
            **{f: fitur_ini[f] for f in FITUR_NON_LAG},
            **lag_bulan_ini,
        })

        # ── OUTPUT 2: PREDIKSI BULAN DEPAN ───────────────────
        bulan_depan, tahun_depan = bulan_berikutnya(bulan, tahun)

        # Carry-forward assumption:
        # Fitur non-lag bulan depan = nilai aktual bulan ini
        # (estimasi terbaik yang tersedia)
        fitur_depan_nonlag = fitur_ini.copy()

        # Lag bulan depan = nilai aktual bulan ini
        # (carry-forward: bulan ini menjadi "bulan lalu" untuk bulan depan)
        lag_bulan_depan = {
            lag_col: fitur_ini[src_col]
            for lag_col, src_col in LAG_SOURCE_MAP.items()
        }

        fitur_prediksi = {**fitur_depan_nonlag, **lag_bulan_depan}
        label_prediksi, prob_prediksi = prediksi_dengan_model(
            artifact, fitur_prediksi
        )

        hasil_prediksi.append({
            'Nama Vendor'     : vendor,
            'Bulan Prediksi'  : bulan_depan,
            'Tahun Prediksi'  : tahun_depan,
            'Prediksi Risiko' : label_prediksi,
            'Prob Patuh'      : prob_prediksi['Patuh'],
            'Prob Waspada'    : prob_prediksi['Waspada'],
            'Prob Kritis'     : prob_prediksi['Kritis'],
            'Probabilitas'    : (
                f"Patuh={prob_prediksi['Patuh']:.0%} | "
                f"Waspada={prob_prediksi['Waspada']:.0%} | "
                f"Kritis={prob_prediksi['Kritis']:.0%}"
            ),
            'Rekomendasi'     : REKOMENDASI.get(label_prediksi, '-'),
            'Asumsi'          : (
                f"Fitur non-lag Feb = nilai aktual {bulan} "
                f"(carry-forward) | Lag Feb = nilai aktual {bulan}"
            ),
        })

    return pd.DataFrame(hasil_monitoring), pd.DataFrame(hasil_prediksi)


# ═══════════════════════════════════════════════════════════════
# FUNGSI — SIMPAN KE EXCEL
# ═══════════════════════════════════════════════════════════════

def simpan_excel(df_mon, df_pred, bulan, tahun,
                 model_name, cv_f1):
    """Simpan dua output ke satu file Excel dengan sheet terpisah."""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    bulan_depan, tahun_depan = bulan_berikutnya(bulan, tahun)
    filename = (OUTPUT_DIR /
                f"hasil_{bulan}_{tahun}_{timestamp}.xlsx")

    wb = Workbook()
    s  = Side(style='thin', color='BFBFBF')
    bd = Border(left=s, right=s, top=s, bottom=s)

    def hdr(cell, bg='1F4E79', fc='FFFFFF', size=11):
        cell.fill      = PatternFill('solid', start_color=bg)
        cell.font      = Font(bold=True, color=fc, size=size)
        cell.alignment = Alignment(horizontal='center',
                                   vertical='center',
                                   wrap_text=True)
        cell.border    = bd

    def row_style(cell, bold=False, size=11, align='center',
                  wrap=False):
        cell.font      = Font(bold=bold, size=size)
        cell.alignment = Alignment(horizontal=align,
                                   vertical='center',
                                   wrap_text=wrap)
        cell.border    = bd

    def warna_kelas(cell, label, bold=True, size=11):
        w = WARNA.get(label, {'bg': 'FFFFFF', 'font': '000000'})
        cell.fill = PatternFill('solid', start_color=w['bg'])
        cell.font = Font(bold=bold, size=size, color=w['font'])
        cell.alignment = Alignment(horizontal='center',
                                   vertical='center')
        cell.border = bd

    def tulis_judul(ws, teks, subjudul, merge_range):
        ws.merge_cells(merge_range)
        start = merge_range.split(':')[0]
        ws[start] = teks
        ws[start].font      = Font(bold=True, size=14,
                                   color='1F4E79')
        ws[start].alignment = Alignment(horizontal='center',
                                        vertical='center')
        ws.row_dimensions[1].height = 32

        r2 = merge_range.replace('1', '2')
        ws.merge_cells(r2)
        s2 = r2.split(':')[0]
        ws[s2] = subjudul
        ws[s2].font      = Font(italic=True, size=10,
                                color='595959')
        ws[s2].alignment = Alignment(horizontal='center')
        ws.row_dimensions[2].height = 18
        ws.row_dimensions[3].height = 8

    def tulis_ringkasan(ws, df_hasil, kolom_label, start_row):
        ws.merge_cells(f'A{start_row}:G{start_row}')
        ws[f'A{start_row}']      = 'RINGKASAN'
        ws[f'A{start_row}'].font = Font(bold=True, size=11,
                                        color='1F4E79')
        ws[f'A{start_row}'].alignment = Alignment(vertical='center')

        for label in ['Patuh', 'Waspada', 'Kritis']:
            start_row += 1
            n   = (df_hasil[kolom_label] == label).sum()
            pct = n / len(df_hasil) * 100
            vendors = df_hasil[
                df_hasil[kolom_label] == label
            ]['Nama Vendor'].tolist()
            ws.merge_cells(f'A{start_row}:G{start_row}')
            c       = ws[f'A{start_row}']
            c.value = (f"  {label}: {n} vendor "
                       f"({pct:.0f}%) — "
                       f"{', '.join(vendors) if vendors else '-'}")
            c.fill  = PatternFill('solid',
                                  start_color=WARNA[label]['bg'])
            c.font  = Font(bold=True, size=11,
                           color=WARNA[label]['font'])
            c.alignment = Alignment(vertical='center')

    # ── SHEET 1: MONITORING BULAN INI ───────────────────────
    ws1 = wb.active
    ws1.title = f'Monitoring {bulan} {tahun}'

    tulis_judul(
        ws1,
        f'MONITORING RISIKO VENDOR — {bulan} {tahun}',
        (f"Klasifikasi aktual berdasarkan data input  |  "
         f"Model: {model_name}  |  "
         f"Dibuat: {datetime.now().strftime('%d/%m/%Y %H:%M')}"),
        'A1:G1'
    )

    # Penjelasan dua kolom klasifikasi
    ws1.merge_cells('A3:G3')
    ws1['A3'] = (
        'Catatan: Klasifikasi Formula = dari skor risiko expert judgment '
        '| Klasifikasi Model = dari Random Forest (untuk validasi silang)'
    )
    ws1['A3'].font      = Font(italic=True, size=9, color='595959')
    ws1['A3'].alignment = Alignment(horizontal='center',
                                    vertical='center')
    ws1.row_dimensions[3].height = 16

    hdrs1 = ['No', 'Nama Vendor', 'Bulan', 'Tahun',
             'Skor Risiko', 'Klasifikasi\n(Formula)',
             'Klasifikasi\n(Model RF)', 'Probabilitas Model',
             'Rekomendasi Tindakan']
    for col, h in enumerate(hdrs1, 1):
        hdr(ws1.cell(row=4, column=col, value=h))
    ws1.row_dimensions[4].height = 38

    for i, (_, row) in enumerate(df_mon.iterrows(), 1):
        r     = i + 4
        label = row['Klasifikasi (Formula)']
        lmodel = row['Klasifikasi (Model)']

        vals = [i, row['Nama Vendor'], row['Bulan'],
                row['Tahun'], row['Skor Risiko'],
                label, lmodel,
                row['Probabilitas'], row['Rekomendasi']]

        for col, val in enumerate(vals, 1):
            cell = ws1.cell(row=r, column=col, value=val)
            row_style(cell)
            if col == 5:   # skor risiko
                cell.number_format = '0.00'
            elif col == 6: # klasifikasi formula
                warna_kelas(cell, label)
            elif col == 7: # klasifikasi model
                warna_kelas(cell, lmodel)
            elif col == 9: # rekomendasi
                row_style(cell, align='left', wrap=True)
        ws1.row_dimensions[r].height = 26

    tulis_ringkasan(ws1, df_mon, 'Klasifikasi (Formula)',
                    len(df_mon) + 6)

    for col, w in enumerate([4, 16, 10, 7, 10, 14, 14, 42, 42], 1):
        ws1.column_dimensions[get_column_letter(col)].width = w

    # ── SHEET 2: PREDIKSI BULAN DEPAN ───────────────────────
    ws2 = wb.create_sheet(f'Prediksi {bulan_depan} {tahun_depan}')

    tulis_judul(
        ws2,
        f'PREDIKSI RISIKO VENDOR — {bulan_depan} {tahun_depan}',
        (f"Prediksi menggunakan carry-forward assumption dari data "
         f"{bulan} {tahun}  |  Model: {model_name}  |  "
         f"CV F1: {cv_f1:.4f}"),
        'A1:G1'
    )

    # Penjelasan carry-forward
    ws2.merge_cells('A3:G3')
    ws2['A3'] = (
        f"Metodologi: Fitur non-lag {bulan_depan} diestimasi "
        f"dari nilai aktual {bulan} {tahun} "
        f"(carry-forward assumption — LOCF). "
        f"Lag {bulan_depan} = nilai aktual {bulan} {tahun}."
    )
    ws2['A3'].font      = Font(italic=True, size=9, color='595959')
    ws2['A3'].alignment = Alignment(horizontal='center',
                                    vertical='center',
                                    wrap_text=True)
    ws2.row_dimensions[3].height = 28

    hdrs2 = ['No', 'Nama Vendor', 'Bulan Prediksi',
             'Tahun Prediksi', 'Prediksi Risiko',
             'Probabilitas', 'Rekomendasi Tindakan']
    for col, h in enumerate(hdrs2, 1):
        hdr(ws2.cell(row=4, column=col, value=h), bg='2E4057')
    ws2.row_dimensions[4].height = 38

    for i, (_, row) in enumerate(df_pred.iterrows(), 1):
        r     = i + 4
        label = row['Prediksi Risiko']

        vals = [i, row['Nama Vendor'],
                row['Bulan Prediksi'], row['Tahun Prediksi'],
                label, row['Probabilitas'], row['Rekomendasi']]

        for col, val in enumerate(vals, 1):
            cell = ws2.cell(row=r, column=col, value=val)
            row_style(cell)
            if col == 5:
                warna_kelas(cell, label)
            elif col == 7:
                row_style(cell, align='left', wrap=True)
        ws2.row_dimensions[r].height = 26

    tulis_ringkasan(ws2, df_pred, 'Prediksi Risiko',
                    len(df_pred) + 6)

    for col, w in enumerate([4, 16, 14, 13, 16, 42, 42], 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # ── SHEET 3: Detail Fitur ────────────────────────────────
    ws3 = wb.create_sheet('Detail Fitur')

    h3 = (['No', 'Vendor', 'Bulan', 'Tahun',
           'Klasifikasi Formula', 'Klasifikasi Model'] +
          FITUR_NON_LAG + FITUR_LAG + ['Status Lag'])
    for col, h in enumerate(h3, 1):
        hdr(ws3.cell(row=1, column=col, value=h), bg='374151')
    ws3.row_dimensions[1].height = 36

    for i, (_, row) in enumerate(df_mon.iterrows(), 1):
        r      = i + 1
        label  = row['Klasifikasi (Formula)']
        lmodel = row['Klasifikasi (Model)']
        vals   = ([i, row['Nama Vendor'], row['Bulan'],
                   row['Tahun'], label, lmodel] +
                  [row.get(f, 0.0) for f in FITUR_NON_LAG] +
                  [row.get(f, 0.0) for f in FITUR_LAG] +
                  [row['Status Lag']])

        for col, val in enumerate(vals, 1):
            cell = ws3.cell(row=r, column=col, value=val)
            row_style(cell, size=10)
            if col == 5:
                warna_kelas(cell, label, size=10)
            elif col == 6:
                warna_kelas(cell, lmodel, size=10)
            elif col > 6 + len(FITUR_NON_LAG):
                cell.fill = PatternFill('solid',
                                        start_color='EBF3FB')
        ws3.row_dimensions[r].height = 20

    ws3.column_dimensions['A'].width = 4
    ws3.column_dimensions['B'].width = 14
    ws3.column_dimensions['C'].width = 10
    ws3.column_dimensions['D'].width = 7
    ws3.column_dimensions['E'].width = 16
    ws3.column_dimensions['F'].width = 16
    for col in range(7, len(h3) + 1):
        ws3.column_dimensions[get_column_letter(col)].width = 18

    wb.save(filename)
    return filename


# ═══════════════════════════════════════════════════════════════
# FUNGSI — BUAT TEMPLATE INPUT
# ═══════════════════════════════════════════════════════════════

def buat_template():
    fname = 'Data_Input_Bulan_Ini.xlsx'
    if Path(fname).exists():
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = 'Input Data'
    s  = Side(style='thin', color='BFBFBF')
    bd = Border(left=s, right=s, top=s, bottom=s)

    headers = (['Nama Vendor', 'Bulan Tagihan', 'Tahun Tagihan'] +
               FITUR_NON_LAG)

    for col, h in enumerate(headers, 1):
        cell           = ws.cell(row=1, column=col, value=h)
        cell.fill      = PatternFill('solid', start_color='1F4E79')
        cell.font      = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center',
                                   vertical='center',
                                   wrap_text=True)
        cell.border    = bd
    ws.row_dimensions[1].height = 36

    for i, vendor in enumerate(VENDORS_AKTIF, 2):
        ws.cell(row=i, column=1, value=vendor).border = bd
        ws.cell(row=i, column=2, value='JANUARI').border = bd
        ws.cell(row=i, column=3, value=2026).border = bd
        for col in range(4, len(headers) + 1):
            cell           = ws.cell(row=i, column=col, value=0.0)
            cell.fill      = PatternFill('solid',
                                         start_color='FFF2CC')
            cell.alignment = Alignment(horizontal='center',
                                       vertical='center')
            cell.border    = bd
        ws.row_dimensions[i].height = 22

    r_note = len(VENDORS_AKTIF) + 3
    ws.merge_cells(f'A{r_note}:J{r_note}')
    ws[f'A{r_note}'] = (
        'Isi sel KUNING dengan data aktual bulan ini. '
        'Ubah Bulan Tagihan dan Tahun Tagihan sesuai periode. '
        'Script akan menghasilkan: '
        '(1) Klasifikasi risiko bulan ini, dan '
        '(2) Prediksi risiko bulan depan.'
    )
    ws[f'A{r_note}'].font = Font(italic=True, color='9C5700',
                                  size=10)

    widths = [16, 16, 14] + [18] * len(FITUR_NON_LAG)
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    wb.save(fname)
    return True


# ═══════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════

if __name__ == '__main__':

    print("=" * 62)
    print("  KLASIFIKASI & PREDIKSI RISIKO VENDOR BULANAN")
    print("  Model: Random Forest Eks.4 (lag + class weight)")
    print("=" * 62)

    # 1. Load model
    print("\n[1] Memuat model...")
    artifact   = load_model()
    model_name = artifact['model_name']
    cv_f1      = artifact['cv_f1']
    print(f"  Model '{model_name}' berhasil dimuat")
    print(f"  CV Macro F1={cv_f1:.4f} | "
          f"Test Macro F1={artifact['test_f1']:.4f}")

    # 2. Load historis
    print("\n[2] Memuat data historis...")
    df_history = load_history()
    print(f"  {len(df_history)} baris historis dimuat")

    # 3. Buat template jika belum ada
    print("\n[3] Memeriksa template input...")
    if buat_template():
        print("  Template baru dibuat: Data_Input_Bulan_Ini.xlsx")
        print("  Silakan isi data vendor lalu jalankan ulang script")
        exit()
    print("  Template sudah ada")

    # 4. Load dan bersihkan input
    INPUT_FILE = 'Data_Input_Bulan_Ini.xlsx'
    print(f"\n[4] Membaca data input: {INPUT_FILE}")

    if not Path(INPUT_FILE).exists():
        print(f"  File tidak ditemukan: {INPUT_FILE}")
        exit()

    df_input = pd.read_excel(INPUT_FILE)

    # Bersihkan baris kosong dan vendor tidak aktif
    df_input = df_input.dropna(
        subset=['Nama Vendor', 'Tahun Tagihan']
    )
    df_input = df_input[
        df_input['Nama Vendor'].astype(str).str.strip() != ''
    ]
    df_input = df_input[
        df_input['Nama Vendor'].str.strip().isin(VENDORS_AKTIF)
    ].reset_index(drop=True)

    bulan = str(df_input['Bulan Tagihan'].iloc[0]).strip().upper()
    tahun = int(df_input['Tahun Tagihan'].iloc[0])
    bulan_dep, tahun_dep = bulan_berikutnya(bulan, tahun)

    print(f"  Periode input : {bulan} {tahun}")
    print(f"  Vendor        : {len(df_input)} vendor")
    print(f"  Akan prediksi : {bulan_dep} {tahun_dep}")

    # 5. Proses dua output
    print("\n[5] Memproses klasifikasi & prediksi...")
    df_mon, df_pred = proses_semua(
        artifact, df_history, df_input
    )

    # 6. Tampilkan ringkasan di terminal
    print()
    print(f"  OUTPUT 1 — KLASIFIKASI {bulan} {tahun}")
    print(f"  {'Vendor':15s} {'Formula':12s} {'Model RF':12s} "
          f"Probabilitas")
    print("  " + "-" * 58)
    for _, row in df_mon.iterrows():
        print(f"  {row['Nama Vendor']:15s} "
              f"{row['Klasifikasi (Formula)']:12s} "
              f"{row['Klasifikasi (Model)']:12s} "
              f"{row['Probabilitas']}")

    print()
    print(f"  OUTPUT 2 — PREDIKSI {bulan_dep} {tahun_dep}")
    print(f"  {'Vendor':15s} {'Prediksi':12s} Probabilitas")
    print("  " + "-" * 58)
    for _, row in df_pred.iterrows():
        print(f"  {row['Nama Vendor']:15s} "
              f"{row['Prediksi Risiko']:12s} "
              f"{row['Probabilitas']}")

    # 7. Simpan Excel
    print("\n[6] Menyimpan hasil ke Excel...")
    output_file = simpan_excel(
        df_mon, df_pred, bulan, tahun, model_name, cv_f1
    )
    print(f"  Tersimpan: {output_file}")

    # 8. Ringkasan akhir
    print()
    print("=" * 62)
    print(f"  RINGKASAN — {bulan} {tahun}")
    print("=" * 62)
    print(f"  Klasifikasi {bulan}:")
    for label in ['Patuh', 'Waspada', 'Kritis']:
        n = (df_mon['Klasifikasi (Formula)'] == label).sum()
        if n > 0:
            vendors = df_mon[
                df_mon['Klasifikasi (Formula)'] == label
            ]['Nama Vendor'].tolist()
            print(f"    {label:10s}: {', '.join(vendors)}")

    print(f"\n  Prediksi {bulan_dep} {tahun_dep}:")
    for label in ['Patuh', 'Waspada', 'Kritis']:
        n = (df_pred['Prediksi Risiko'] == label).sum()
        if n > 0:
            vendors = df_pred[
                df_pred['Prediksi Risiko'] == label
            ]['Nama Vendor'].tolist()
            print(f"    {label:10s}: {', '.join(vendors)}")

    print()
    print(f"  File Excel: {output_file}")
    print("  Sheet 1: Monitoring bulan ini")
    print("  Sheet 2: Prediksi bulan depan")
    print("  Sheet 3: Detail fitur input + lag")
    print("=" * 62)
    print("  Selesai! File siap diserahkan ke manajemen.")
    print("=" * 62)
